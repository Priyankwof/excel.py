from openpyxl import load_workbook
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime

filepath = "C:\\Users\\USER\\Documents\\Availity Automation.xlsx"

wb = load_workbook(filepath)
sheet = wb.active

d5 = sheet.cell(row=4, column= 4)
Npi = d5.value
print(Npi)

d6 = sheet.cell(row=4, column= 5)
patientid = d6.value
print(patientid)

d7 = sheet.cell(row=4, column= 6)
patlnam = d7.value
print(patlnam)

d8 = sheet.cell(row=4, column= 7)
patfnam= d8.value
print(patfnam)

d9 = sheet.cell(row=4, column= 8)
oldformat = str(d9.value)
datetimeobject = datetime.strptime(oldformat,'%Y-%m-%d %H:%M:%S').date()
patdob = datetimeobject.strftime('%m/%d/%Y')
print(patdob)

d10 = sheet.cell(row=4, column= 9)
oldformat = str(d10.value)
datetimeobject = datetime.strptime(oldformat,'%Y-%m-%d %H:%M:%S').date()
startdat = datetimeobject.strftime('%m/%d/%Y')
print(startdat)

d11 = sheet.cell(row=4, column= 10)
oldformat = str(d11.value)
datetimeobject = datetime.strptime(oldformat,'%Y-%m-%d %H:%M:%S').date()
enddat = datetimeobject.strftime('%m/%d/%Y')
print(enddat)



options = Options()
options.binary_location = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
driver = webdriver.Chrome(chrome_options=options, executable_path=r"C:\Users\USER\Desktop\Trial\selenium\chromedriver.exe",)
usernm="ARZION01"
passwd="Arzion@123"
driver.get("https://apps.availity.com/availity/web/public.elegant.login");

elm = driver.find_element_by_id('userId')
elm.clear()
elm.send_keys(usernm)
elm  = driver.find_element_by_id('password')
elm.clear()
elm.send_keys(passwd)
loginbtn = driver.find_element_by_id('loginFormSubmit')
loginbtn.click()

driver.maximize_window()
time.sleep(5)

# auth = driver.find_element_by_xpath('//*[@id="root"]/div[2]/div[2]/div/div/div[2]/div/div/form/fieldset/div[1]/div[2]/div/label')
# auth.click()

# conbtn = driver.find_element_by_xpath('//*[@id="root"]/div[2]/div[2]/div/div/div[2]/div/div/form/button')
# conbtn.click()
# time.sleep(10)

# vrfybtn = driver.find_element_by_xpath('//*[@id="root"]/div[2]/div[2]/div/div/div[2]/div/div/form/div[2]/div/button')
# vrfybtn.click()

# btn = driver.find_element_by_xpath('//*[@id="root"]/div[2]/div[2]/div/div/div[2]/div/div[3]/button[4]')
# btn.click()

claimdropdown = driver.find_element_by_xpath('//*[@id="availity-secondary-navbar-collapse"]/ul[1]/li[2]/a')
claimdropdown.click()
time.sleep(5)

claimstatus = driver.find_element_by_xpath('/html/body/navigation/div/bottom-nav/nav/div/div[2]/ul[1]/li[2]/ul/li/div/div/div[1]/div/div/div/div[1]/div/div/div/a/span/div/div[2]')
claimstatus.click()     
print("Loged In ")

time.sleep(5)
driver.get('https://apps.availity.com/public/apps/home/#!/loadApp?appUrl=%2Fweb%2Fclmsmgmt%2Fclaim-status-ui%2F%3FcacheBust%3D1583263269%23%2Fdashboard%3ForgId%3D10822771%26payerId%3D040')
iframe = driver.find_element_by_id('newBodyFrame')
driver.switch_to.frame(iframe)
print("switched to iframe")

time.sleep(3)
npi = driver.find_element_by_id('providerNpi')
driver.execute_script("document.getElementById('providerNpi').value = '';")
npi.send_keys(Npi)


patid = driver.find_element_by_xpath('//*[@id="patientMemberId"]')
patid.send_keys(patientid)

pat_firstname = driver.find_element_by_id('patientFirstName')
pat_firstname.send_keys(patfnam)

pat_lastname = driver.find_element_by_id('patientLastName')
pat_lastname.send_keys(patlnam)

DateofBirth = driver.find_element_by_id('patientBirthDate')
DateofBirth.send_keys(patdob)

Start_date = driver.find_element_by_css_selector('#serviceDates-start')
Start_date.click()
Start_date.send_keys(startdat)

End_date = driver.find_element_by_css_selector('#serviceDates-end')
End_date.click()
End_date.clear()
End_date.send_keys(enddat)

submit = driver.find_element_by_id('submit-by276')
submit.click()
print('Claim submitted ')

time.sleep(8)
print('Claim Report found')
time.sleep(3)
h2s = driver.find_element_by_css_selector('.mb-3.card.card-body')
claim = h2s.find_elements_by_css_selector('h2.my-2')

print('Verifying Type:', type(h2s))

for h2 in claim:
    
    if 'Claim' in h2.text:
        print(h2.text)
    else:
         print('Element Not found')

span = h2s.find_elements_by_css_selector('.span')
spanclass = h2s.find_elements_by_css_selector('span.font-weight-bold')

for date in spanclass:

    if 'Dates of Service' in date.text:
        print(date.text)
    else:
        print('dos not found')

for billed in spanclass:

    if 'Billed' in  billed.text:
        print(billed.text)
    else:
        print('billed not found')

for denied in spanclass:

    if 'DENIED' in denied.text:
        print(denied.text)
    else:
        print('denied not found')

for paid in spanclass:

    if 'Paid'in paid.text:
        print(paid.text)
    else:
        print('Paid not found')


#lists = h2.find_element_by_class_name('span.text-muted')
#print(lists.text)

# claimno = driver.find_element_by_class_name('h2.my-2')
# print ('claimno'+ claimno.text)

# span = driver.find_elements_by_class_name('span.font-weight-bold')
# print(span)
# billed = driver.find_element_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(2) > div:nth-child(2) > div > div.bg-faded.p-2.border.col-sm-6 > div > div:nth-child(2) > span')
# print('billed'+ billed.text)

# paid = driver.find_element_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(2) > div:nth-child(2) > div > div.bg-dark.p-2.col-sm-6 > div > div:nth-child(2) > span')
# print('paid'+ paid.text)

# processdate = driver.find_element_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(2) > div:nth-child(1) > div > div:nth-child(2) > div:nth-child(2) > span')
# print('processdate'+ processdate.text)

# status = driver.find_element_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(2) > div:nth-child(1) > div > div:nth-child(3) > div:nth-child(2) > span')
# # status = driver.find_elements_by_class_name('mb-1 col-sm-auto col-lg-auto')
# # status.find`
# print('status'+ status.text)

# CHECKNUM = driver.find_elements_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(5) > div > div > div:nth-child(1) > div:nth-child(2) > span')
# print('CHECKNUM' + CHECKNUM.text)

# Pataccount = driver.find_element_by_css_selector('#pageContainer > div:nth-child(5) > div.col > div > div > div:nth-child(5) > div > div > div:nth-child(2) > div:nth-child(2) > span')
# print('Pataccount' + Pataccount.text)

